from django.contrib.auth import forms
from django.http import response
from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from pandas._libs import missing
from .decorators import unauthenticated_user, allowed_users
from django.contrib.auth.models import Group
from openpyxl import load_workbook
import os
import shutil
from oletools.olevba import VBA_Parser, TYPE_OLE, TYPE_OpenXML, TYPE_Word2003_XML, TYPE_MHTML
import re
import subprocess
from pathlib import Path
from matplotlib import pyplot as plt
import urllib, base64
import io
import pandas as pd


filename = ""

@login_required(login_url='/registration/login/')
def login_request(request):
    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.info(request, f"You are now logged in as {username}")
            return render(request, "index.html")

    form = AuthenticationForm()
    return render(request = request,
                    template_name = "login.html",
                    context={"form":form})

def home(request):
    return render(request, 'index.html')

@unauthenticated_user
def login(request):
    """
    Creates login view
    Returns: rendered login page
    """
    return render(request, 'login.html')


def register(response):
    if response.method == 'POST':
        form = UserCreationForm(response.POST)
        if form.is_valid():
            user = form.save()           
            if response.POST.get('group') == 'administrator':
                group = Group.objects.get(name='New user')
            user.groups.add(group)
            return redirect("login")
    else:
        form = UserCreationForm()
    return render(response, 'registration/register.html', {'form': form})

def logout_request(request):
    logout(request)
    messages.info(request, "you have successfully logged out.")
    return redirect('login')

@allowed_users(allowed_roles=["administrator", "New user"])
def upload(request):
    result = []
    hidden_data = []
    incomplete_data = []
    macro = []
    if request.method == 'POST':
        if request.FILES.get('document'):
            file = request.FILES['document']

            EXCEL_FILE_EXTENSIONS = ('xlsb', 'xls', 'xlsm', 'xla', 'xlt', 'xlam',)
            KEEP_NAME = False  # Set this to True if you would like to keep "Attribute VB_Name"
            
            workbook = load_workbook(filename=file, data_only=True)
            xls = workbook[workbook.sheetnames[0]] 

            # macros check

            count = 0
            
            # check .exe, url, missing data,  
            for row in xls.iter_rows(min_row=1, max_col=xls.max_column, max_row=xls.max_row):
                count = count + 1
                for column_value in row:
                    null_data_check = column_value.value

                    if re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', str(null_data_check)):
                        message = 'found a url link in  row ', count
                        result.append(message)

                    elif re.findall(r'\b\S*\.exe\b', str(null_data_check)):
                        message = 'found .exe file in  row ', count
                        result.append(message)

            
            # check hidden rows
            for rowLetter,rowDimension in xls.row_dimensions.items():
                if rowDimension.hidden == True:
                    hidden_row = str('found hidden rows at ', rowLetter)
                    hidden_data.append(hidden_row)

            if not hidden_data:
                hidden_data.append('Hidden data not found')

            if not result:
                result.append(".exe, excel formula and, url aren't found")

            # missing data report 
            df = pd.read_excel(file)
            missing_data = df.isnull().sum()
            incomplete_data = str(missing_data).split('\n')

            
            filedata = open(file.name, 'rb').read()
            vbaparser = VBA_Parser(file.name, data=filedata)

            if vbaparser.detect_vba_macros():
                macro.append("Caution, macros has been found in your excel file.")

            if not macro:
                macro.append("No macro found")
    
        dic_result = {'data': incomplete_data, 'hidden': hidden_data, 'result': result, 'macro': macro}
        return render(request, 'report.html', dic_result)
    return render(request, 'upload.html')

@allowed_users(allowed_roles=["administrator", "New user"])
def extract(request):
    result = []
    hidden_data = []
    incomplete_data = []
    extract_macro = []
    number = []
    input_wrong = ["check out your input data and follow the how to input section"]
    
    if request.method == 'POST':
        if request.FILES.get('document'):
            file = request.FILES['document']
            data = request.POST.get('data')
            
            print(request.POST.get('data'))
            workbook = load_workbook(filename=file, data_only=True)
            xls = workbook[workbook.sheetnames[0]] 

            # macros check

            count = 0
            
            

            if data == 'hidden':
                # check hidden rows
                for rowLetter,rowDimension in xls.row_dimensions.items():
                    if rowDimension.hidden == True:
                        hidden_data.append(rowDimension)

                if not hidden_data:
                    hidden_data.append('Hidden data not found')

                dic_result = {'data': hidden_data}
                return render(request, 'extract_data.html', dic_result)

            elif str(data) == 'missing':
                # missing data report 
                df = pd.read_excel(file)
                missing_data = df.isnull().sum()
                incomplete_data = str(missing_data).split('\n')
                dic_result = {'data': incomplete_data}           
                return render(request, 'extract_data.html', dic_result)
            
            elif str(data) == "macro":
                filedata = open(file.name, 'rb').read()
                vbaparser = VBA_Parser(file.name, data=filedata)

                for (filename, stream_path, vba_filename, vba_code) in vbaparser.extract_macros(): 
                    extract_macro.append(str(vba_code))
                
                if not extract_macro:
                    extract_macro.append("No macro found")

                dic_result = {'data': extract_macro}
                return render(request, 'extract_data.html', dic_result)

            # check .exe, url, missing data,
            elif str(data) == "url":
                for row in xls.iter_rows(min_row=1, max_col=xls.max_column, max_row=xls.max_row):
                    count = count + 1
                    for column_value in row:
                        null_data_check = column_value.value

                        if re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', str(null_data_check)):
                            result.append(str(null_data_check))

                if not result:
                    result.append("Url data aren't found")
                dic_result = {'data': result}
            
                return render(request, 'extract_data.html', dic_result)

            elif str(data) == ".exe":
                for row in xls.iter_rows(min_row=1, max_col=xls.max_column, max_row=xls.max_row):
                    count = count + 1
                    for column_value in row:
                        null_data_check = column_value.value
                        if re.findall(r'\b\S*\.exe\b', str(null_data_check)):
                            result.append(str(null_data_check))
                
                if not result:
                    result.append(".exe data aren't found")
                dic_result = {'data': result}
            
                return render(request, 'extract_data.html', dic_result)

            elif isinstance(float(data), float):
                for row in xls.iter_rows(min_row=1, max_col=xls.max_column, max_row=xls.max_row):
                    count = count + 1
                    for column_value in row:
                        data_check = column_value.value
                        if str(data) == str(data_check):
                            number.append(data_check)
                dic_result = {'data': number}
            
                return render(request, 'extract_data.html', dic_result)


        dic_result = {'data': input_wrong}
            
        return render(request, 'extract.html', dic_result)
                
                
    return render(request, 'extract.html')